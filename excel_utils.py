from __future__ import annotations

import io
import os
import re
import tempfile
import unicodedata
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Iterable

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


SIEFORES = [
    "SIEFORE 95",
    "SIEFORE 60",
    "SIEFORE 65",
    "SIEFORE 70",
    "SIEFORE 75",
    "SIEFORE 80",
    "SIEFORE 85",
    "SIEFORE 90",
    "SIEFORE INICIAL",
]


@dataclass
class ProcessResult:
    path: Path
    messages: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def deaccent(value: object) -> str:
    text = "" if value is None else str(value)
    return "".join(
        ch for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )


def normalize_siefore(series: pd.Series) -> pd.Series:
    return (
        series.astype(str).str.strip().map(deaccent).str.upper()
        .str.replace(r"^INVER60.*$", "SIEFORE 60", regex=True)
        .str.replace(r"^INVER65.*$", "SIEFORE 65", regex=True)
        .str.replace(r"^INVER70.*$", "SIEFORE 70", regex=True)
        .str.replace(r"^INVER75.*$", "SIEFORE 75", regex=True)
        .str.replace(r"^INVER80.*$", "SIEFORE 80", regex=True)
        .str.replace(r"^INVER85.*$", "SIEFORE 85", regex=True)
        .str.replace(r"^INVER90.*$", "SIEFORE 90", regex=True)
        .str.replace(r"^INVER95.*$", "SIEFORE 95", regex=True)
        .str.replace(r"^INVERIN.*$", "SIEFORE INICIAL", regex=True)
    )


def save_upload(uploaded_file, folder: Path, name: str | None = None) -> Path | None:
    if uploaded_file is None:
        return None
    folder.mkdir(parents=True, exist_ok=True)
    suffix = Path(uploaded_file.name).suffix
    target = folder / (name or uploaded_file.name)
    if not target.suffix and suffix:
        target = target.with_suffix(suffix)
    target.write_bytes(uploaded_file.getbuffer())
    return target


def save_uploads(uploaded_files, folder: Path) -> list[Path]:
    if not uploaded_files:
        return []
    return [p for f in uploaded_files if (p := save_upload(f, folder)) is not None]


def make_workdir() -> Path:
    return Path(tempfile.mkdtemp(prefix="derivados_streamlit_"))


def clear_cells(ws, min_row: int, max_row: int, columns: Iterable[int]) -> None:
    for row in range(min_row, max_row + 1):
        for col in columns:
            ws.cell(row=row, column=col).value = None


def clear_used_columns(path: Path, sheet: str, min_row: int, columns: Iterable[int]) -> None:
    wb = load_workbook(path)
    ws = wb[sheet]
    clear_cells(ws, min_row, ws.max_row, columns)
    wb.save(path)
    wb.close()


def write_dataframe(
    path: Path,
    sheet_name: str,
    df: pd.DataFrame,
    startrow: int,
    startcol: int = 0,
    header: bool = True,
) -> None:
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=startrow,
            startcol=startcol,
            index=False,
            header=header,
        )


def parse_money(value: object) -> float:
    text = str(value).strip()
    if not text or text.lower() == "nan" or text == "-":
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace("$", "").replace("\xa0", "").replace(" ", "")
    if re.search(r",\d{1,2}$", text):
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return -number if negative else number


def fetch_fix_value(timeout: int = 12) -> float:
    response = requests.get("https://www.banxico.org.mx/tipcamb/tipCamMIAction.do", timeout=timeout)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    for td in soup.find_all("td"):
        text = td.get_text(strip=True)
        if text.replace(".", "").isdigit():
            return float(text)
    raise ValueError("No se encontro el tipo de cambio FIX en Banxico.")


def try_recalculate_excel(path: Path) -> bool:
    if os.name != "nt":
        return False
    try:
        import win32com.client as win32  # type: ignore
    except Exception:
        return False
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(path))
        excel.CalculateFull()
        workbook.Save()
        workbook.Close(SaveChanges=True)
        return True
    except Exception:
        return False
    finally:
        if excel is not None:
            excel.Quit()


def workbook_bytes(path: Path) -> bytes:
    return path.read_bytes()
