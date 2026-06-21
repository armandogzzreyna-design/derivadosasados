from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .excel_utils import (
    ProcessResult,
    SIEFORES,
    clear_used_columns,
    deaccent,
    normalize_siefore,
    try_recalculate_excel,
    write_dataframe,
)


def days_360(start: date, end: date, method: str = "US") -> int:
    sy, sm, sd = start.year, start.month, start.day
    ey, em, ed = end.year, end.month, end.day
    if method.upper() == "US":
        if sd == 31:
            sd = 30
        if ed == 31 and sd in (30, 31):
            ed = 30
    else:
        if sd == 31:
            sd = 30
        if ed == 31:
            ed = 30
    return (ey - sy) * 360 + (em - sm) * 30 + (ed - sd)


def parse_fecha_code(value: object) -> date | None:
    if value is None:
        return None
    match = re.search(r"(\d{6})\s*$", str(value).strip())
    if not match:
        return None
    yy, mm, dd = match.group(1)[:2], match.group(1)[2:4], match.group(1)[4:6]
    try:
        return date(int("20" + yy), int(mm), int(dd))
    except ValueError:
        return None


def calc_haircut(counterparty: str, dias: int) -> float:
    cpt = counterparty.strip().upper()
    if cpt == "BNP":
        if 1 <= dias <= 360:
            return 0.995
        if 361 <= dias <= 720:
            return 0.98
        if 721 <= dias <= 1800:
            return 0.97
        if 1801 <= dias <= 3600:
            return 0.95
        if dias >= 3601:
            return 0.92
    if "MORGAN" in cpt or "GOLDMAN" in cpt:
        if 1 <= dias <= 360:
            return 0.995
        if 361 <= dias <= 1800:
            return 0.98
        if dias >= 1801:
            return 0.96
    if "BBVA" in cpt:
        if 1 <= dias <= 364:
            return 0.99
        if 365 <= dias <= 1092:
            return 0.98
        if 1093 <= dias <= 2548:
            return 0.97
        if 2549 <= dias <= 3639:
            return 0.96
        if dias >= 3640:
            return 0.94
    return 1.0


def paste_txt_into_vectores(vectores_path: Path, deuda_path: Path | None, derivados_path: Path | None) -> bool:
    wb = load_workbook(vectores_path)
    if deuda_path:
        ws = wb["DEUDA"]
        for i, line in enumerate(deuda_path.read_text(encoding="latin-1", errors="replace").splitlines(), start=2):
            ws.cell(row=i, column=1, value=line)
    if derivados_path:
        ws = wb["DERIVADOS"]
        for i, line in enumerate(derivados_path.read_text(encoding="latin-1", errors="replace").splitlines(), start=2):
            ws.cell(row=i, column=1, value=line)
    wb.save(vectores_path)
    wb.close()
    return try_recalculate_excel(vectores_path)


def update_vectores_back(path_otc: Path, vectores_path: Path) -> int:
    df_deuda = pd.read_excel(vectores_path, sheet_name="DEUDA", usecols=["EMISION_V2", "PRECIO SUCIO", "ISIN"])
    df_der = pd.read_excel(vectores_path, sheet_name="DERIVADOS", usecols=["EMISION_V2", "PRECIO SUCIO"])
    filtros_deuda = ["TBON", "TNOT", "M BONOS 280302", "UDIBONO 401115", "M BONOS 290301"]
    filtros_derivados = ["FWD SPX"]
    pat_deuda = r"(?:" + "|".join(re.escape(p) for p in filtros_deuda) + r")"
    pat_der = r"(?:" + "|".join(re.escape(p) for p in filtros_derivados) + r")"
    filtered = pd.concat(
        [
            df_deuda[df_deuda["EMISION_V2"].astype(str).str.contains(pat_deuda, case=False, na=False, regex=True)],
            df_der[df_der["EMISION_V2"].astype(str).str.contains(pat_der, case=False, na=False, regex=True)],
        ],
        ignore_index=True,
    )
    clear_used_columns(path_otc, "Inst & Gar", 4, [1, 2, 3])
    write_dataframe(path_otc, "Inst & Gar", filtered, startrow=3)
    return len(filtered)


def write_fix(path_otc: Path, fix_value: float) -> None:
    wb = load_workbook(path_otc)
    wb["Inst & Gar"]["C2"] = float(fix_value)
    wb.save(path_otc)
    wb.close()


def update_haircuts(path_otc: Path, method: str = "US") -> int:
    wb_values = load_workbook(path_otc, data_only=True)
    ws_values = wb_values["Directorio"]
    wb_write = load_workbook(path_otc, data_only=False, keep_vba=path_otc.suffix.lower() == ".xlsm")
    ws_write = wb_write["Directorio"]
    ws_write["G1"] = ws_write["G1"].value or "Resultado"
    ws_write["J1"] = ws_write["J1"].value or f"Dias_30_360_{method}"
    ws_write["K1"] = ws_write["K1"].value or "HAIRCUT"
    count = 0
    for row in range(2, ws_values.max_row + 1):
        fecha = parse_fecha_code(ws_values[f"C{row}"].value)
        if not fecha:
            ws_write[f"G{row}"] = None
            ws_write[f"J{row}"] = None
            ws_write[f"K{row}"] = None
            continue
        dias = abs(days_360(fecha, date.today(), method=method))
        contraparte = str(ws_values[f"B{row}"].value or "")
        ws_write[f"J{row}"] = dias
        ws_write[f"K{row}"] = calc_haircut(contraparte, dias)
        count += 1
    wb_values.close()
    wb_write.save(path_otc)
    wb_write.close()
    return count


def append_cash_movements(path_otc: Path, movimientos_cash: Path | None) -> int:
    if not movimientos_cash:
        return 0
    df = pd.read_excel(movimientos_cash, sheet_name=0)
    df = df.drop(columns=[c for c in ["Td Num", "Counterparty", "Tran Type"] if c in df.columns])
    df = df.rename(columns={"Trade Date": "FECHA", "Fund": "SIEFORE", "Principal": "CANTIDAD"})
    df["CANTIDAD"] = pd.to_numeric(df["CANTIDAD"], errors="coerce").fillna(0).abs()
    if len(df) > 0:
        df = df.drop(df.index[-1])
    df["SIEFORE"] = normalize_siefore(df["SIEFORE"])
    grouped = df.groupby(["MOVIMIENTO", "CONTRAPARTE", "FECHA", "SIEFORE"], dropna=False)["CANTIDAD"].sum().unstack("SIEFORE", fill_value=0).reset_index()
    headers = ["MOVIMIENTO", "CONTRAPARTE", "FECHA", *SIEFORES]
    for col in headers:
        if col not in grouped.columns:
            grouped[col] = 0 if col.startswith("SIEFORE") else pd.NA
    aligned = grouped[headers]
    wb = load_workbook(path_otc)
    start_row = wb["CASH"].max_row
    wb.close()
    write_dataframe(path_otc, "CASH", aligned, startrow=start_row, header=False)
    return len(aligned)


def append_otc_movements(path_otc: Path, movimientos_otc: Path | None) -> int:
    if not movimientos_otc:
        return 0
    df = pd.read_excel(movimientos_otc, sheet_name=0)
    df = df.rename(columns={
        "Counterparty": "CONTRAPARTE",
        "Trade Date": "FECHA",
        "Fund": "SIEFORE",
        "IVC": "INSTRUMENTO",
        "Quantity": "TITULOS",
        "Tran Type": "MOVIMIENTO",
    })
    df = df.drop(columns=[c for c in ["Td Num"] if c in df.columns])
    df["MOVIMIENTO"] = (
        df["MOVIMIENTO"].astype(str).str.strip().map(deaccent).str.upper()
        .str.replace(r"^CIN RETURN\b.*", "ENVIO", regex=True)
        .str.replace(r"^CIN\b.*", "RECEP", regex=True)
    )
    df["INSTRUMENTO"] = df["INSTRUMENTO"].astype(str).str[3:]
    if len(df) > 0:
        df = df.drop(df.index[-1])
    df["TITULOS"] = pd.to_numeric(df["TITULOS"], errors="coerce").fillna(0).abs() / 100
    df["CONTRAPARTE"] = (
        df["CONTRAPARTE"].astype(str).str.strip().map(deaccent).str.upper()
        .str.replace(r"^CBGSMX.*$", "GOLDMAN", regex=True)
        .str.replace(r"^DBBNP.*$", "BNP", regex=True)
        .str.replace(r"^DMSPLC.*$", "MORGAN", regex=True)
    )
    df["SIEFORE"] = normalize_siefore(df["SIEFORE"])
    grouped = df.groupby(["CONTRAPARTE", "FECHA", "INSTRUMENTO", "MOVIMIENTO", "SIEFORE"], dropna=False)["TITULOS"].sum().unstack("SIEFORE", fill_value=0).reset_index()
    headers = ["MOVIMIENTO", "CONTRAPARTE", "INSTRUMENTO", "FECHA", *SIEFORES]
    for col in headers:
        if col not in grouped.columns:
            grouped[col] = 0 if col.startswith("SIEFORE") else pd.NA
    aligned = grouped[headers]
    wb = load_workbook(path_otc)
    start_row = wb["Movimientos"].max_row
    wb.close()
    write_dataframe(path_otc, "Movimientos", aligned, startrow=start_row, header=False)
    return len(aligned)


def update_totales(path_otc: Path) -> int:
    df = pd.read_excel(path_otc, sheet_name="Movimientos").drop(columns=["FECHA"], errors="ignore")
    mov = df["MOVIMIENTO"].astype(str).str.strip().str.upper()
    sign = np.where(mov == "ENVIO", -1, 1)
    for col in SIEFORES:
        df[col] = pd.to_numeric(df.get(col, 0), errors="coerce").fillna(0) * sign
    totals = df.groupby(["CONTRAPARTE", "INSTRUMENTO"], as_index=False)[SIEFORES].sum()
    totals = totals.loc[~(totals[SIEFORES] == 0).all(axis=1)].copy()
    totals[SIEFORES] = totals[SIEFORES].astype(object)
    totals.loc[:, SIEFORES] = totals[SIEFORES].mask(np.isclose(totals[SIEFORES].astype(float), 0.0), "")
    clear_used_columns(path_otc, "Totales", 2, range(1, 12))
    write_dataframe(path_otc, "Totales", totals, startrow=0)
    return len(totals)


def run_otc_process(
    path_otc: Path,
    fix_value: float,
    vectores_path: Path | None = None,
    deuda_path: Path | None = None,
    derivados_path: Path | None = None,
    movimientos_otc: Path | None = None,
    movimientos_cash: Path | None = None,
) -> ProcessResult:
    messages: list[str] = []
    warnings: list[str] = []
    if vectores_path:
        recalculated = paste_txt_into_vectores(vectores_path, deuda_path, derivados_path) if (deuda_path or derivados_path) else False
        rows = update_vectores_back(path_otc, vectores_path)
        messages.append(f"Vectores OTC actualizados: {rows} filas.")
        if (deuda_path or derivados_path) and not recalculated:
            warnings.append("No se pudo recalcular VECTORES con Excel en este sistema; se usaron los valores guardados en el archivo.")
    write_fix(path_otc, fix_value)
    messages.append(f"FIX escrito en Inst & Gar: {fix_value}.")
    messages.append(f"Haircuts calculados: {update_haircuts(path_otc)} filas.")
    messages.append(f"Movimientos CASH agregados: {append_cash_movements(path_otc, movimientos_cash)} filas.")
    messages.append(f"Movimientos OTC agregados: {append_otc_movements(path_otc, movimientos_otc)} filas.")
    messages.append(f"Totales recalculados: {update_totales(path_otc)} filas.")
    return ProcessResult(path=path_otc, messages=messages, warnings=warnings)
