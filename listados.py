from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .excel_utils import ProcessResult, clear_used_columns, parse_money, try_recalculate_excel, write_dataframe


def copy_previous_day(path_listados: Path) -> None:
    wb = load_workbook(path_listados)
    wb_values = load_workbook(path_listados, data_only=True)
    src = wb_values["VALIDADOR REP_SALDOS (EDITAR)"]
    dst = wb["ARCHIVO_DIA_ANTERIOR"]
    for i, row in enumerate(src.iter_rows(min_row=1, max_row=56, min_col=1, max_col=14, values_only=True), start=1):
        for j, value in enumerate(row, start=1):
            dst.cell(row=i, column=j).value = value
    wb.save(path_listados)
    wb.close()
    wb_values.close()


def update_vector_precios(path_listados: Path, path_vector: Path | None) -> int:
    if not path_vector:
        return 0
    wb = load_workbook(path_listados)
    ws = wb["LLAMADAS DE MARGEN"]
    for row in range(2, 11):
        ws.cell(row=row, column=7).value = None
        ws.cell(row=row, column=10).value = None
    for row in range(24, 33):
        ws.cell(row=row, column=7).value = None
        ws.cell(row=row, column=10).value = None
    for row in range(46, 57):
        ws.cell(row=row, column=7).value = None
        ws.cell(row=row, column=11).value = None
    wb.save(path_listados)
    wb.close()

    df_vector = pd.read_excel(path_vector, skiprows=4)
    result = df_vector[df_vector["Serie"].isin([401115, 340824])][["Serie", "Precio Descontado"]]
    wb = load_workbook(path_listados)
    ws = wb["FUT&GAR MEX"]
    for row in [6, 7]:
        raw = ws[f"A{row}"].value
        if not raw:
            continue
        serie = str(raw).split()[-1]
        precio = result.loc[result["Serie"] == int(serie), "Precio Descontado"].values
        if len(precio) > 0:
            ws[f"B{row}"] = precio[0]
    wb.save(path_listados)
    wb.close()
    return len(result)


def update_llamadas_retiros(path_listados: Path, path_llamadas: Path | None) -> int:
    if not path_llamadas:
        return 0
    df = pd.read_excel(path_llamadas, sheet_name="llamadas|diferencias")
    df["Monto"] = df["Monto"].map(parse_money)
    wb = load_workbook(path_listados)
    ws = wb["VALIDADOR REP_SALDOS (EDITAR)"]
    for row in range(2, 57):
        ws[f"E{row}"] = 0
        ws[f"F{row}"] = 0
        ws[f"G{row}"] = 0
        ws[f"H{row}"] = 0
    count = 0
    for _, record in df.iterrows():
        portfolio = str(record["Portafolio"]).strip().lower()
        monto = float(record["Monto"])
        for row in range(2, ws.max_row + 1):
            current = str(ws[f"A{row}"].value).strip().lower()
            if portfolio == current:
                if monto > 0:
                    ws[f"F{row}"] = -abs(monto)
                elif monto < 0:
                    ws[f"E{row}"] = abs(monto)
                count += 1
                break
    wb.save(path_listados)
    wb.close()
    return count


def _clean_number_series(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace("\xa0", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace("-", "0", regex=False)
    )
    return pd.to_numeric(cleaned, errors="coerce").fillna(0)


def update_int_com(path_listados: Path, path_int_com: Path | None) -> int:
    if not path_int_com:
        return 0
    df_diario = pd.read_excel(path_int_com, sheet_name="Scotia", header=None)
    tabla = pd.DataFrame({
        "Portafolio": df_diario.iloc[0, 1:].tolist(),
        "Intereses": df_diario.iloc[3, 1:].tolist(),
        "Comision": df_diario.iloc[12, 1:].tolist(),
    }).replace("-", 0)
    src_wb = load_workbook(path_int_com, data_only=True)
    dst_wb = load_workbook(path_listados)
    src_ws = src_wb["Scotia"]
    dst_ws = dst_wb["INT & COM"]
    for i in range(8, 16):
        for j in range(2, 12):
            value = src_ws.cell(row=i, column=j).value
            dst_ws.cell(row=152 + (i - 8), column=j + 1, value=0 if value == "-" else value)
    dst_wb.save(path_listados)
    dst_wb.close()
    src_wb.close()

    tabla["Intereses"] = _clean_number_series(tabla["Intereses"])
    tabla["Comision"] = _clean_number_series(tabla["Comision"])

    def code(portfolio: object) -> str:
        text = str(portfolio).upper()
        if "IN" in text:
            return "BCSCOTINF"
        digits = "".join(ch for ch in text if ch.isdigit())[:2]
        return f"BCSCOT{digits}F"

    tabla["Codigo"] = tabla["Portafolio"].apply(code)
    resumen = tabla.groupby("Codigo", as_index=False)[["Comision", "Intereses"]].sum()
    tabla_final = pd.read_excel(path_int_com, sheet_name="Tabla final")[["Portafolio", "Comisiones", "Intereses"]]
    tabla_final = tabla_final.rename(columns={"Comisiones": "Comision", "Portafolio": "Codigo"})
    base = pd.concat([resumen, tabla_final], ignore_index=True)

    wb = load_workbook(path_listados)
    ws = wb["VALIDADOR REP_SALDOS (EDITAR)"]
    for row in range(2, ws.max_row + 1):
        ws[f"G{row}"] = 0
        ws[f"H{row}"] = 0
    count = 0
    for _, record in base.iterrows():
        codigo = str(record["Codigo"]).strip().lower()
        for row in range(2, ws.max_row + 1):
            current = str(ws[f"A{row}"].value or "").strip().lower()
            if codigo == current:
                ws[f"G{row}"] = abs(float(record["Intereses"]))
                ws[f"H{row}"] = -abs(float(record["Comision"]))
                count += 1
                break
    wb.save(path_listados)
    wb.close()
    return count


def write_fix(path_listados: Path, fix_value: float) -> None:
    wb = load_workbook(path_listados)
    wb["CME"]["B1"] = float(fix_value)
    wb.save(path_listados)
    wb.close()


def update_vectores_listados(path_listados: Path, vectores_path: Path | None) -> int:
    if not vectores_path:
        return 0
    df_deuda = pd.read_excel(vectores_path, sheet_name="DEUDA", usecols=["EMISION_V2", "PRECIO SUCIO", "ISIN"])
    year = datetime.now().year
    last_digit = year % 10
    last_two = year % 100
    cme = [f"{letter}{last_digit}" for letter in ["H", "Z", "U", "M"]]
    deuas = [f"DEUA {month}{last_two}" for month in ["MR", "JN", "SP", "DC"]]
    pat_cme = r"(?:" + "|".join(re.escape(p) for p in cme) + r")$"
    pat_deuas = r"(?:" + "|".join(re.escape(p) for p in deuas) + r")"
    emision = df_deuda["EMISION_V2"].astype(str).str.strip()
    filtered = pd.concat(
        [
            df_deuda[emision.str.contains(pat_cme, case=False, na=False, regex=True) & ~emision.str.contains(r"SP", case=False, na=False)],
            df_deuda[emision.str.contains(pat_deuas, case=False, na=False, regex=True)],
        ],
        ignore_index=True,
    )
    clear_used_columns(path_listados, "Vectores", 1, [1, 2, 3])
    write_dataframe(path_listados, "Vectores", filtered, startrow=0)
    return len(filtered)


def update_aims(path_listados: Path, aims_path: Path | None) -> tuple[int, bool]:
    if not aims_path:
        return 0, False
    wb = load_workbook(path_listados)
    ws = wb["GARANTIAS "]
    for row in range(66, 100):
        for col in [15, 16, 17, 18, 19, 22, 23]:
            ws.cell(row=row, column=col).value = None
    wb.save(path_listados)
    wb.close()

    src_wb = load_workbook(aims_path, data_only=True)
    dst_wb = load_workbook(path_listados)
    src_ws = src_wb["DetalleAIM"]
    dst_ws = dst_wb["GARANTIAS "]
    for i in range(2, 30):
        for j in range(1, 6):
            value = src_ws.cell(row=i, column=j).value
            dst_ws.cell(row=64 + i, column=j + 14, value=0 if value == "-" else value)
        for j in range(6, 9):
            value = src_ws.cell(row=i, column=j).value
            dst_ws.cell(row=64 + i, column=j + 16, value=0 if value == "-" else value)
    dst_wb.save(path_listados)
    dst_wb.close()
    recalculated = try_recalculate_excel(path_listados)

    dst_wb = load_workbook(path_listados, data_only=True)
    dst_ws = dst_wb["GARANTIAS "]
    for i in range(65, 95):
        for j in range(15, 25):
            src_ws.cell(row=i - 64, column=j - 14, value=dst_ws.cell(row=i, column=j).value)
    src_wb.save(aims_path)
    src_wb.close()
    dst_wb.close()
    return 30, recalculated


def paste_estados_de_cuenta(path_listados: Path, estados_df: pd.DataFrame | None) -> int:
    if estados_df is None or estados_df.empty:
        return 0
    wb = load_workbook(path_listados)
    ws = wb["ESTADOS_DE_CUENTA"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    wb.save(path_listados)
    wb.close()
    write_dataframe(path_listados, "ESTADOS_DE_CUENTA", estados_df, startrow=1, header=False)
    return len(estados_df)


def run_listados_process(
    path_listados: Path,
    fix_value: float,
    path_vector: Path | None = None,
    path_llamadas: Path | None = None,
    path_int_com: Path | None = None,
    path_vectores: Path | None = None,
    path_aims: Path | None = None,
    estados_df: pd.DataFrame | None = None,
) -> ProcessResult:
    messages: list[str] = []
    warnings: list[str] = []
    copy_previous_day(path_listados)
    messages.append("Copia del dia anterior actualizada.")
    messages.append(f"Vector de precios actualizado: {update_vector_precios(path_listados, path_vector)} filas.")
    messages.append(f"Llamadas/retiros pegados: {update_llamadas_retiros(path_listados, path_llamadas)} coincidencias.")
    messages.append(f"Intereses y comisiones pegados: {update_int_com(path_listados, path_int_com)} coincidencias.")
    write_fix(path_listados, fix_value)
    messages.append(f"FIX escrito en CME: {fix_value}.")
    messages.append(f"Vectores Listados actualizados: {update_vectores_listados(path_listados, path_vectores)} filas.")
    aims_rows, recalculated = update_aims(path_listados, path_aims)
    messages.append(f"AIMS actualizado: {aims_rows} filas.")
    if aims_rows and not recalculated:
        warnings.append("AIMS requiere formulas de Excel; en este sistema se usaron los valores calculados que ya tenia la plantilla.")
    messages.append(f"Estados de cuenta pegados: {paste_estados_de_cuenta(path_listados, estados_df)} filas.")
    return ProcessResult(path=path_listados, messages=messages, warnings=warnings)
